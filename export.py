import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import stats
# os.chdir("natural_selection_sim-main/")

# Vide le fichier excel
def clear(): 
    workbook = openpyxl.load_workbook("simulations.xlsx")
    for e in workbook: # On supprime toutes les pages
        workbook.remove(e)
    workbook.create_sheet(title='simulation') # Recrée une page
    workbook.save("simulations.xlsx")

# Crée une nouvelle feuille de calcul dans le fichier
def create_sheet():
    workbook = openpyxl.load_workbook("simulations.xlsx")
    workbook.create_sheet(title='simulation')
    workbook.save("simulations.xlsx")

# Ecrit l'intitulé des colonnes 
def load_and_write():
    workbook = openpyxl.load_workbook("simulations.xlsx")
    name_page = workbook.sheetnames[-1] # On récupère le nom de la dernière feuille
    ws = workbook[name_page] # On se place dans la feuille en question
    ws.append(["Génération", "Nombre d'individus départ", "Nombre d'individus en vie", "Nombre de naissances", "Nombre de morts", "Moyenne taille individus", "Moyenne vue individus", "Moyenne vitesse individus", "Nombre de morts total", ])
    workbook.save("simulations.xlsx")

# Exporte les données de la génération actuelle
def export():
    workbook = openpyxl.load_workbook("simulations.xlsx")
    name_page = workbook.sheetnames[-1] # On récupère le nom de la dernière feuille
    ws = workbook[name_page] # On se place dans la feuille en question
    ws.append([stats.generation, stats.nb_individus_start, stats.nb_individus_alive, stats.births, stats.nb_individus_dead, stats.individus_moyenne_size, stats.individus_moyenne_view, stats.individus_moyenne_speed, stats.nb_individus_dead_total])
    # print([stats.generation, stats.nb_individus_start, stats.nb_individus_alive, stats.births, stats.nb_individus_dead, stats.individus_moyenne_size, stats.individus_moyenne_view, stats.individus_moyenne_speed, stats.nb_individus_dead_total])
    workbook.save("simulations.xlsx")

# Test pour la réalisation du graphique
# def graph():
#     workbook = openpyxl.load_workbook("simulations.xlsx")
#     name_page = workbook.sheetnames[-1] # On récupère le nom de la dernière feuille
#     ws = workbook[name_page]
    
#     values = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=stats.generation+1)
#     graphique = BarChart()
#     graphique.title = "Nombre d'individus dans le temps"
#     graphique.x_axis_title = "Generation"
#     graphique.y_axis_title = "Nombre d'individus vivants"
#     graphique.add_data(values, "A10")
#     ws.add_chart(graphique)
#     workbook.save("simulations.xlsx")
    

    
    
