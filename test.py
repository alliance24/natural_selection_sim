# # import openpyxl

# # data_frame = pd.read_excel('simulations.xlsx')

# import os, stats
# os.chdir("natural_selection_sim-main/")
# import pandas as pd

# workbook = pd.read_excel("simulations.xlsx")

# # Créer un DataFrame avec une ligne de données
# data = [{"Génération": stats.generation, "Nombre d'individus départ": stats.nb_individus_start, "Nombre d'individus en vie": stats.nb_individus_alive, "Nombre de naissances": stats.births, "Nombre de morts": stats.nb_individus_dead, "Moyenne taille individus": stats.individus_moyenne_size, "Moyenne vue individus": stats.individus_moyenne_view, "Moyenne vitesse individus": stats.individus_moyenne_speed, "Nombre de morts total": stats.nb_individus_dead_total}]
# data_frame_to_export = pd.DataFrame.from_records(data)
# data_frame_to_export.to_excel("simulations.xlsl")
import openpyxl, os, stats
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
os.chdir("natural_selection_sim-main/")

# workbook = openpyxl.load_workbook("simulations.xlsx")
# name_page = workbook.sheetnames[-1] # On récupère le nom de la dernière feuille
# ws = workbook[name_page]
# graphique = BarChart()
# graphique.title = "Nombre d'individus dans le temps"
# graphique.x_axis_title = "Generation"
# graphique.y_axis_title = "Nombre d'individus vivants"
# valeurs_x = Reference(ws, min_col=1, min_row=2, max_row=stats.generation+1)
# valeurs_y = Reference(ws, min_col=3, min_row=2, max_row=stats.generation+1)
# serie = openpyxl.chart.Series(valeurs_y, xvalues=valeurs_x, title="Données")
# graphique.append(serie)
# ligne = stats.generation+1
# colonne = 1
# lettre_colonne = get_column_letter(colonne)
# cellule = ws[lettre_colonne + str(ligne)]
# ws.add_chart(graphique, cellule)
# workbook.save("simulations.xlsx")

workbook = openpyxl.load_workbook("simulations.xlsx")
for e in workbook: # On supprime toutes les pages
    workbook.remove(e)
workbook.create_sheet(title='simulation') # Recrée une page
workbook.save("simulations.xlsx")